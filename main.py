import requests
import logging
import argparse
import openpyxl




MOSCOW_REGION = 113
SVERDL_REGION = 1261
MONTH_PERIOD = 30
MOSCOW_REGION_SJ = 4
VERBOSITY_TO_LOGGING_LEVELS = {
    0: logging.WARNING,
    1: logging.INFO,
    2: logging.DEBUG,
}


def get_predict_salary(sallaries):
    predicted_salary_list = []
    for salary in sallaries:
        if salary['from'] is not None and salary['to'] is not None:
            predicted_salary = (salary['from'] + salary['to']) / 2
        elif salary['from'] is not None and salary['to'] is None:
            predicted_salary = salary['from'] * 1.2
        elif salary['from'] is None and salary['to'] is not None:
            predicted_salary = salary['to'] * 0.8
        else:
            continue
        predicted_salary_list.append(predicted_salary)
    if predicted_salary_list:
        processed_salaries = len(predicted_salary_list)
        average_salary = sum(predicted_salary_list) // len(
            predicted_salary_list)
    else:
        processed_salaries, average_salary = (0, 0)
    return processed_salaries, average_salary


def get_vacancies_hh_keywords(language, region, period=None):
    url = "https://api.hh.ru/vacancies"
    params = {
        "area": region,
        "text": language,
        "period": period,
        "page": 0,
        # "areas": 1261
    }
    response = requests.get(url, params=params)
    response_json = response.json()
    founded_vanacies = response_json['found']
    vacancies = response_json['items']
    pages = response_json['pages']
    for page in range(1, pages):
        logging.info(f'Loading  {language} page : {page}....')
        params["page"] = page
        sallaries = requests.get(url, params=params)
        vacancies.extend(sallaries.json()['items'])
    return founded_vanacies, vacancies


def get_vacancies_sj(language, region, period=None):
    secret_key = "v3.r.128977974.a1abb3a24d8881daebd4b8268c0bb09839c95ca9.6c1f43473acc583d623876c00b4f6f535ede39dd"
    url = "https://api.superjob.ru/2.0/vacancies/"
    headers = {
        "X-Api-App-Id": secret_key
    }
    params = {
        "keyword": language,
        "town": region,
    }
    response = requests.get(url, headers=headers, params=params)
    response_json = response.json()
    founded_vacancies = response_json['total']
    vacancies = response_json['objects']
    return founded_vacancies, vacancies


def get_only_rub_av_salary_hh(vacancies):
    all_salaries = []
    for item in vacancies:
        if not item or not item['salary']:
            continue
        if item['salary']['currency'] != 'RUR':
            continue
        all_salaries.append(item['salary'])
    processed_salaries, average_salary = get_predict_salary(all_salaries)
    return processed_salaries, average_salary


def get_only_rub_av_salary_sj(vacancies):
    all_salaries = []
    for item in vacancies:
        # унифицируем ключи
        item["from"] = item.pop("payment_from")
        item["to"] = item.pop("payment_to")
        all_salaries.append(item)
    processed_salaries, average_salary = get_predict_salary(all_salaries)
    return processed_salaries, average_salary


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--verbose', '-v', action='count', default=1)
    args = parser.parse_args()
    logging_level = VERBOSITY_TO_LOGGING_LEVELS[args.verbose]
    logging.basicConfig(level=logging_level)
    # top_lang_list = ["Программист Python", "Программист Ruby",
    #                   "Программист Java",
    #                   "Программист Swift", "Программист Javascript",
    #                   "Программист Go",
    #                   "Программист C++", "Программист PHP", "Программист C#"]
    top_lang_list = ["Разработчик python junior"]

    wb = openpyxl.load_workbook(filename="./test.xlsx")
    sheet_msk = wb['vacancies_msk']
    sheet_sverdl = wb['vacancies_sverdl']


    # сначала ищем по ключевым словам мск
# {'id': '30264584', 'premium': False, 'name': 'DevOps инженер', 'department': None, 'has_test': False, 'response_letter_required': False, 'area': {'id': '3', 'name': 'Екатеринбург', 'url': 'https://api.hh.ru/areas/3'}, 'salary': None, 'type': {'id': 'open', 'name': 'Открытая'}, 'address': {'city': 'Екатеринбург', 'street': None, 'building': None, 'description': None, 'lat': 56.838607, 'lng': 60.605514, 'raw': None, 'metro': {'station_name': 'Площадь 1905 года', 'line_name': 'Север-Юг', 'station_id': '48.266', 'line_id': '48', 'lat': 56.837982, 'lng': 60.59734}, 'metro_stations': [{'station_name': 'Площадь 1905 года', 'line_name': 'Север-Юг', 'station_id': '48.266', 'line_id': '48', 'lat': 56.837982, 'lng': 60.59734}], 'id': '800454'}, 'response_url': None, 'sort_point_distance': None, 'employer': {'id': '78638', 'name': 'Тинькофф', 'url': 'https://api.hh.ru/employers/78638', 'alternate_url': 'https://hh.ru/employer/78638', 'logo_urls': {'240': 'https://hhcdn.ru/employer-logo/2848221.png', '90': 'https://hhcdn.ru/employer-logo/2848220.png', 'original': 'https://hhcdn.ru/employer-logo-original/601766.png'}, 'vacancies_url': 'https://api.hh.ru/vacancies?employer_id=78638', 'trusted': True}, 'published_at': '2019-06-03T11:17:55+0300', 'created_at': '2019-06-03T11:17:55+0300', 'archived': False, 'apply_alternate_url': 'https://hh.ru/applicant/vacancy_response?vacancyId=30264584', 'insider_interview': None, 'url': 'https://api.hh.ru/vacancies/30264584?host=hh.ru', 'alternate_url': 'https://hh.ru/vacancy/30264584', 'relations': [], 'snippet': {'requirement': 'Способности автоматизировать рутинные процессы на <highlighttext>Python</highlighttext>, bash. Отличное знание сети. Возможность рассказать про модель OSI, умение диагностировать и решать сложные...', 'responsibility': 'Поддерживать команды <highlighttext>Python</highlighttext> <highlighttext>разработчиков</highlighttext> в вопросах инфраструктурной оптимизации и сопровождения. Развертывать и поддерживать CI/CD на основе TeamCity и Jenkins. '}, 'contacts': None}
    for language in top_lang_list:
        hh_founded_vacancies_quantity, msk_hh_vacancies = get_vacancies_hh_keywords(language, MOSCOW_REGION, MONTH_PERIOD)

        r = 1
        for vacancy in msk_hh_vacancies:
            print(vacancy)
            sheet_msk.cell(row=r, column=1, value=" ".join(language.split()[0:2]))
            sheet_msk.cell(row=r, column=2, value=language.split()[-1])
            if vacancy['salary'] is not None:
                sheet_msk.cell(row=r, column=3, value=vacancy['salary']['from'])
                sheet_msk.cell(row=r, column=4, value=vacancy['salary']['to'])
            sheet_msk.cell(row=r, column=5, value=vacancy['name'])

            r += 1

    wb.save("./test.xlsx")



if __name__ == '__main__':
    main()